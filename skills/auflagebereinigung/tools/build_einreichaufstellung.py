# -*- coding: utf-8 -*-
"""Einreich-Aufstellung (Formulare / Eingaben) Auflagebereinigung 2619-KISPI.

Gegenstück zur Plan-/Dokumentenliste (build_auflagenliste.py): nicht nach Planer,
sondern nach EINGABE-ADRESSAT (Amt/Stelle) gegliedert. Beantwortet: welches amtliche
Formular / welche Unterlage geht an welches Amt, bis wann, und WO bekomme ich das
Formular (Spalte Bezugsquelle, klickbar). Grundlage: definitiver Bauentscheid
BE 1171/26 vom 08.06.2026 (Geschäfts-Nr. B26-00705.01), Stadt Zürich.
Bezugsquellen am 13.06.2026 live verifiziert (HTTP 200) bzw. wörtlich aus dem BE.
JANS-Layout (neutral, schwarz)."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

# ------------------------------------------------------------------ Gruppen (Amt)
# Jede Gruppe: (amt_lang, kanal_hinweis, [zeilen]).
# Zeile: (formular, auflage, frist, (bezugsquelle_text, url_oder_None), status)
GROUPS = [
("UGZ – Umwelt- und Gesundheitsschutz Zürich, Energie im Bau",
 "Eingabekanal: stadt-zuerich.ch/ugz-baubewilligung  ·  Energienachweise verbindlich über EVEN (energievollzug.ch), seit 01.01.2026",
 [
  ("Energienachweis – Hauptformular EN-ZH (Deckblatt)", "Ziff. II.8a / II.1c",
   "4 Wo vor Baubeginn (ca. 03.07.2026)",
   ("EVEN-Plattform energievollzug.ch (Arbeitsfassung EN-ZH-005)", "https://www.energievollzug.ch"),
   "ausgefüllt (EN-ZH-005, 13.06.) · EVEN-Erfassung offen"),
  ("Energienachweis – Lüftungstechnische Anlagen (EN-105)", "Ziff. II.8b",
   "4 Wo vor Baubeginn", ("EVEN-Plattform energievollzug.ch", "https://www.energievollzug.ch"),
   "Entwurf bei HLK Gruner (Kontrolle bis 15.06.)"),
  ("Energienachweis – Kühlung / Befeuchtung (EN-110-ZH)", "Ziff. II.8c",
   "4 Wo vor Baubeginn", ("zh.ch → EN-110-ZH (PDF)",
   "https://www.zh.ch/content/dam/zhweb/bilder-dokumente/themen/planen-bauen/bauvorschriften/bauvorschriften-im-energiebereich/energienachweis/projektkontrolle/en_110_zh.pdf"),
   "Entwurf bei HLK Gruner (Kontrolle bis 15.06.)"),
  ("Nachweis genehmigte Abwärmenutzungsanlage Klima (oder Unwirtschaftlichkeit, §30a BBV I)",
   "Erwägung d) / Ziff. II.8", "4 Wo vor Baubeginn",
   ("UGZ Energie im Bau (stadt-zuerich.ch/ugz-baubewilligung)", "https://www.stadt-zuerich.ch/ugz-baubewilligung"),
   "offen"),
  ("Schallschutznachweis SIA 181 (Aussen-/Trennbauteile, haustechnische Anlagen); ggf. Erleichterungsgesuch schriftlich",
   "Ziff. II.10", "Ausführungsplanung",
   ("UGZ Lärmschutz / NIS (stadt-zuerich.ch/ugz-baubewilligung)", "https://www.stadt-zuerich.ch/ugz-baubewilligung"),
   "offen"),
 ]),
("Feuerpolizei Stadt Zürich (Brandschutz)",
 "Einreichung digital an die Feuerpolizei; Prüfung durch QS-Verantwortliche/n Brandschutz (QSS 3)",
 [
  ("QS-Konzept / Nachweise Qualitätssicherung Brandschutz Stufe QSS 3 (VKF)", "Ziff. II.11–13",
   "vor Baubeginn / laufend", ("VKF-Brandschutzrichtlinie 'Qualitätssicherung' (praever.ch)", "https://www.praever.ch"),
   "offen"),
  ("Zusatzformular 3 Brandschutznachweis (Baueingabe-Beilage)", "Erwägung g) / Ziff. II.16",
   "mit Baueingabe", ("stadt-zuerich.ch → Zusatzformular 3 Brandschutznachweis (PDF)",
   "https://www.stadt-zuerich.ch/content/dam/web/de/planen-bauen/baubewilligungen/dokumente/bewilligungsprozess/dokumente/zusatzformular-3-brandschutznachweis.pdf"),
   "eingereicht (Baueingabe)"),
  ("Überarbeitetes Brandschutzkonzept + überarbeitete Brandschutzpläne (gesamter Umbaubereich)",
   "Ziff. II.16 / II.1b", "vor Baubeginn", ("projektspezifisch – BRA Gruner (Jens Ziegel)", None),
   "offen"),
  ("Detailpläne Schiebetüren (Fluchttür + Brandabschluss) mit VKF-Anerkennung / Leistungserklärung",
   "Ziff. II.18 / II.1b", "vor Baubeginn", ("VKF-Brandschutzregister (bsronline.ch)", "https://www.bsronline.ch"),
   "offen"),
  ("Detailpläne brandschutzrelevante Innenwandkonstruktionen (Material + Aufbau)", "Ziff. II.19 / II.1b",
   "vor Baubeginn", ("projektspezifisch – BRA Gruner", None), "offen"),
  ("Vollständiges Lüftungskonzept lufttechnische Anlagen (Umbaubereich)", "Ziff. II.21 / II.3",
   "vor Arbeitsvergabe", ("projektspezifisch – HLK Gruner (Cekdar Duran)", None), "offen"),
  ("Übereinstimmungserklärung Brandschutz (VKF) + nachgef. BS-/Feuerwehrpläne (1.OG) + BSK-Revision + Integral-Test-Protokolle",
   "Ziff. II.2 / II.24", "vor Inbetriebnahme", ("VKF-Übereinstimmungserklärung (praever.ch)", "https://www.praever.ch"),
   "offen"),
 ]),
("Gebäudeversicherung Kanton Zürich (GVZ)",
 "Projektunterlagen vor Ausführungsbeginn einreichen und genehmigen lassen",
 [
  ("Brandmeldeanlage (BMA) – Projektunterlagen (Weisung 'Brandmeldeanlagen')", "Ziff. II.25 / II.3",
   "vor Arbeitsvergabe", ("GVZ Kanton Zürich (gvz.ch)", "https://www.gvz.ch"),
   "offen · ELE Gruner"),
  ("Sprinkleranlage (SPA) – Projektunterlagen (Weisung 'Sprinkleranlagen')", "Ziff. II.26 / II.3",
   "vor Arbeitsvergabe", ("GVZ Kanton Zürich (gvz.ch)", "https://www.gvz.ch"),
   "offen · SPR JOMOS"),
 ]),
("Amt für Baubewilligungen Stadt Zürich (AfB)",
 "Sammelbestätigungen (Ziff. II.1) und Meldungen; teils Bestätigungen anderer Stellen beizulegen",
 [
  ("Abgeänderte Pläne Fluchtwegführung (II.15) + Austauschpläne Reduit 4.2 – einreichen + bewilligen lassen",
   "Ziff. II.1a / Erwägung i)", "vor Baubeginn", ("projektspezifisch – ARC (Raphael Jans)", None),
   "offen"),
  ("Bestätigung der Feuerpolizei zu II.16 / II.18 / II.19", "Ziff. II.1b", "vor Baubeginn",
   ("von Feuerpolizei einholen, an AfB", None), "offen"),
  ("Bestätigung UGZ Energie im Bau zu II.8 (energetische Anforderungen)", "Ziff. II.1c", "vor Baubeginn",
   ("von UGZ einholen, an AfB", None), "offen"),
  ("Meldung Baubeginn + Bauvollendung (§327 PBG, §23 BVV)", "Ziff. II.31", "4 Wo vor / nach",
   ("Amt für Baubewilligungen Stadt Zürich", None), "offen"),
 ]),
("Schutzraum – Feuerpolizei / Schutzbauten (SRZ)",
 "Schutzraumpflicht abklären; bei Pflicht Eingabeunterlagen EAG-Projekte",
 [
  ("Schutzraum-/EAG-Eingabeunterlagen (Erforderliche Unterlagen Eingabe EAG-Projekte)", "Ziff. II.1b (SRZ)",
   "vor Baubeginn", ("SRZ Schutzbauten (Merkblatt AS-2025-657-DE liegt im Projekt)", None),
   "Schutzraumpflicht in Abklärung"),
 ]),
("Bauinstallation / öffentlicher Grund (4 Wochen vor Baubeginn)",
 "Kontaktaufnahme spätestens 4 Wo vor geplantem Baubeginn",
 [
  ("Anmeldung Bauinstallation / Bauzufahrt bei Benützung Privatgrund", "Ziff. II.28a", "4 Wo vor Baubeginn",
   ("AfB Baukontrolle · Tel. 044 412 28 65 (08.00–09.00)", None), "offen"),
  ("Bewilligung Benützung öffentlicher Grund", "Ziff. II.28b", "4 Wo vor Baubeginn",
   ("Stadtpolizei, Kommissariat Bewilligung Vollzug · Tel. 044 411 72 73", None), "offen"),
  ("Baumerhalt-Abklärung Wurzel-/Kronenbereich (Gesuch/Merkblatt Baumfällung)", "Ziff. II.28c",
   "4 Wo vor Baubeginn", ("Grün Stadt Zürich – gsz-baumerhalt@zuerich.ch / stadt-zuerich.ch (Gesuch Baumfällung)",
   "https://www.stadt-zuerich.ch/content/dam/web/de/planen-bauen/bauberatung-und-dienstleistungen/dokumente/gruenraeume-freiraeume/freiraumberatung/gesuchsformular-und-merkblatt-baumfaellung.pdf"),
   "offen (sofern Bäume betroffen)"),
  ("Gemeinsames Zustandsprotokoll öffentlicher Grund + Festsetzung Depositum", "Ziff. II.29a/b",
   "4 Wo vor Baubeginn", ("Tiefbauamt, Fachbereich Strassen · Tel. 044 412 23 35", None), "offen"),
 ]),
("Baustelle – Umwelt (während Bauzeit)",
 "Vollzug Luftreinhaltung / Bauabfall",
 [
  ("Massnahmen Luftreinhaltung Baustelle (Massnahmenstufe A; Merkblatt)", "Ziff. II.30", "während Bauzeit",
   ("stadt-zuerich.ch/baustellen-luft", "https://www.stadt-zuerich.ch/baustellen-luft"),
   "offen"),
  ("Bauabfall-Trennung + umweltgerechte Entsorgung (VVEA Art. 17)", "Ziff. II.32", "während Bauzeit",
   ("Entsorgungskonzept (liegt vor, Baueingabe)", None), "liegt vor (Baueingabe)"),
 ]),
]

# ------------------------------------------------------------------ Styling
BLACK = "FF000000"
HDR_FILL = PatternFill("solid", fgColor="FF404040")
GRP_FILL = PatternFill("solid", fgColor="FFD9D9D9")
ALT_FILL = PatternFill("solid", fgColor="FFF2F2F2")
TITLE_F  = Font(name="Cambria", size=14, bold=True, color=BLACK)
SUB_F    = Font(name="Cambria", size=10, color=BLACK)
HDR_F    = Font(name="Cambria", size=10, bold=True, color="FFFFFFFF")
GRP_F    = Font(name="Cambria", size=10, bold=True, color=BLACK)
GRPK_F   = Font(name="Cambria", size=9, italic=True, color="FF404040")
CELL_F   = Font(name="Cambria", size=10, color=BLACK)
LINK_F   = Font(name="Cambria", size=10, color="FF0563C1", underline="single")
LEG_H    = Font(name="Cambria", size=10, bold=True, color=BLACK)
thin = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
TOPL = Alignment(vertical="top", horizontal="left", wrap_text=True)
CTR = Alignment(vertical="center", horizontal="left", wrap_text=True)

wb = Workbook()
ws = wb.active
ws.title = "Einreich-Aufstellung"
ws.sheet_view.showGridLines = False
# Druck: A4 quer, feste Skalierung (fitToWidth wird von LibreOffice bei sehr breiten
# Tabellen unzuverlässig umgesetzt -> sonst horizontaler Seitenumbruch). 68 % bringt
# die 7 Spalten sicher auf EINE Seitenbreite.
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=False)
ws.page_setup.fitToWidth = 0
ws.page_setup.fitToHeight = 0
ws.page_setup.scale = 68
ws.page_margins.left = ws.page_margins.right = 0.3
ws.page_margins.top = ws.page_margins.bottom = 0.5

HEAD = ["Nr.", "Eingabe an (Amt/Stelle)", "Amtliches Formular / einzureichende Unterlage",
        "Auflage (Ziff.)", "Frist", "Bezugsquelle / Plattform", "Status"]
WIDTHS = [5, 26, 50, 20, 22, 46, 30]
for i, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
NCOL = len(HEAD)

# Titelblock
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
c = ws.cell(1, 1, "Einreich-Aufstellung Auflagebereinigung – Formulare und Eingaben je Amt")
c.font = TITLE_F; c.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

subs = [
 "Projekt 2619-KISPI «Neue Psychosomatische Therapiestation für Jugendliche» – "
 "Universitäts-Kinderspital Zürich, Eleonorenstiftung, Lenggstrasse 30, 8008 Zürich",
 "Grundlage: definitiver Bauentscheid BE 1171/26 vom 08.06.2026 (Versand 09.06.2026), "
 "Amt für Baubewilligungen Stadt Zürich, Kreisarchitektin J. Palermo  ·  Geschäfts-Nr. B26-00705.01, Code 08303",
 "Grundstück Kat.-Nr. RI5416, Zürich 8 – Weinegg  ·  Gestaltungsplan «Lengg», Wohnanteil 0 %  ·  "
 "kant. Verfügung BVV 26-1211 (Baudirektion Kt. ZH, 18.05.2026, kant. Betrieb, ohne Auflagen)",
 "Stand: 13.06.2026  ·  rechtskräftiger Entscheid; bestätigt den Vorabzug (Ziffern II.1–33, "
 "Energie-Zuordnung II.8a/b/c wörtlich)  ·  geplanter Baubeginn lt. Baugesuch 01.08.2026 → 4-Wochen-Fristen ca. 03.07.2026",
 "Bezugsquellen am 13.06.2026 live geprüft (HTTP 200) bzw. wörtlich aus dem Bauentscheid; Quelltext klickbar in Spalte «Bezugsquelle / Plattform».",
]
r = 2
for s in subs:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    cc = ws.cell(r, 1, s); cc.font = SUB_F; cc.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 26
    r += 1

hdr_row = r + 1
for j, h in enumerate(HEAD, start=1):
    cc = ws.cell(hdr_row, j, h)
    cc.font = HDR_F; cc.fill = HDR_FILL; cc.alignment = CTR; cc.border = BORDER
ws.row_dimensions[hdr_row].height = 18

rr = hdr_row + 1
nr = 0
for amt, kanal, rows in GROUPS:
    # Gruppen-Kopf (Amt) + Kanal-Zeile
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NCOL)
    gc = ws.cell(rr, 1, amt); gc.font = GRP_F; gc.alignment = Alignment(vertical="center", horizontal="left")
    for j in range(1, NCOL + 1):
        ws.cell(rr, j).fill = GRP_FILL; ws.cell(rr, j).border = BORDER
    ws.row_dimensions[rr].height = 16
    rr += 1
    if kanal:
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NCOL)
        kc = ws.cell(rr, 1, kanal); kc.font = GRPK_F; kc.alignment = Alignment(vertical="center", wrap_text=True)
        rr += 1
    for k, (form, aufl, frist, bez, status) in enumerate(rows):
        nr += 1
        bez_text, bez_url = bez
        amt_kurz = amt.split(" –")[0].split(" (")[0]
        vals = [nr, amt_kurz, form, aufl, frist, bez_text, status]
        for j, v in enumerate(vals, start=1):
            cc = ws.cell(rr, j, v)
            cc.border = BORDER
            cc.alignment = Alignment(vertical="top", horizontal="center" if j == 1 else "left", wrap_text=True)
            if j == 6 and bez_url:
                cc.hyperlink = bez_url; cc.font = LINK_F
            else:
                cc.font = CELL_F
            if k % 2 == 1:
                cc.fill = ALT_FILL
        rr += 1
data_last = rr - 1

ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(NCOL)}{data_last}"
ws.freeze_panes = f"A{hdr_row + 1}"
ws.print_title_rows = f"{hdr_row}:{hdr_row}"  # Kopfzeile auf Folgeseiten wiederholen

# ------------------------------------------------------------------ Legenden
rr += 1
def legend_block(title, items):
    global rr
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NCOL)
    cc = ws.cell(rr, 1, title); cc.font = LEG_H
    rr += 1
    for it in items:
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NCOL)
        cc = ws.cell(rr, 1, it); cc.font = CELL_F; cc.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1
    rr += 1

legend_block("Ämter / Plattformen", [
 "UGZ = Umwelt- und Gesundheitsschutz Zürich, Energie im Bau  ·  AfB = Amt für Baubewilligungen Stadt Zürich  ·  "
 "FP = Feuerpolizei Stadt Zürich",
 "GVZ = Gebäudeversicherung Kanton Zürich (gvz.ch)  ·  VKF = Vereinigung Kantonaler Feuerversicherungen (praever.ch; Register bsronline.ch)",
 "EVEN = Elektronischer Vollzug energetischer Nachweise (energievollzug.ch) – seit 01.01.2026 verbindlich für den Energienachweis Kt. ZH",
 "TBA = Tiefbauamt  ·  StaPo = Stadtpolizei Zürich  ·  GSZ = Grün Stadt Zürich  ·  SRZ = Schutz & Rettung Zürich (Schutzbauten)",
])
legend_block("Fristen (Einreichungszeitpunkt gemäss Ziff. II.1–3)", [
 "Ziff. II.1 – vor Baubeginn, an AfB: a) Fluchtwegführung II.15  ·  b) Bestätigung FP zu II.16/II.18/II.19  ·  c) Bestätigung UGZ zu II.8",
 "Ziff. II.2 – vor Inbetriebnahme / feuerpolizeil. Abnahme, an FP: Übereinstimmungserklärung, nachgef. BS-/Feuerwehrpläne, BSK-Revision, Integral-Test-Protokolle",
 "Ziff. II.3 – vor Arbeitsvergabe: II.21 Lüftung  ·  II.25 BMA  ·  II.26 Sprinkler (SPA)",
 "Energienachweise (II.8) mind. 4 Wochen vor Baubeginn an UGZ; geplanter Baubeginn 01.08.2026 → ca. 03.07.2026.",
])
legend_block("Hinweise", [
 "• Diese Aufstellung ergänzt die Plan- und Dokumentenliste (gegliedert nach Planer): hier die Sicht je Amt mit Bezugsquelle der Formulare.",
 "• Energienachweis: seit 01.01.2026 verbindlich über EVEN (energievollzug.ch). EN-110-ZH zusätzlich als PDF auf zh.ch verfügbar; "
 "EN-ZH-Deckblatt + EN-105 werden in EVEN erfasst. Private Kontrolle Energie ist noch zu bestimmen.",
 "• Adress-Hinweis: Der Bauentscheid führt als Zustelladresse «Raphael Jans – Architekten ETC / SIA, Saumstrasse 21, 8003 Zürich». "
 "Aktuelle JANS-Adresse ist Grubenstrasse 37, 8045 Zürich (und «ETH», nicht «ETC») – beim AfB korrigieren lassen, damit Korrespondenz richtig zugestellt wird.",
 "• Status-Werte: offen / in Arbeit / ausgefüllt / eingereicht / genehmigt / erledigt / hinfällig.",
])

# ------------------------------------------------------------------ Speichern
OUT_PROJEKT = ("/Users/raphaeljans/Library/CloudStorage/OneDrive-FreigegebeneBibliotheken–JANS/"
 "JANS - 2619-KISPI - Dokumente/3 Umbauprojekt Neu/33.01_JANS/01 Dokumente Bewilligunsverfahren/"
 "3_Dokumente Auflagebereinigung/00_Checklisten Architekt Plan und Dokumentenliste Komplette Auflagen Insgesamt")
OUT_HUB = ("/Users/raphaeljans/Library/CloudStorage/OneDrive-FreigegebeneBibliotheken–JANS/"
 "AD - 01 Geschaeftsfuerung/JANS AI/30 JANS AI HUB OUTPUT/auflagebereinigung/2619-KISPI")
FNAME = "260613-Einreich-Aufstellung-Formulare-Auflagebereinigung-2619-KISPI.xlsx"

os.makedirs(OUT_PROJEKT, exist_ok=True)
os.makedirs(OUT_HUB, exist_ok=True)
p1 = os.path.join(OUT_PROJEKT, FNAME)
p2 = os.path.join(OUT_HUB, FNAME)
wb.save(p1)
wb.save(p2)
print("Eingaben/Formulare:", nr)
print("gespeichert:", p1)
print("gespeichert:", p2)
