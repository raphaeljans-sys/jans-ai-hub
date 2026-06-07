# -*- coding: utf-8 -*-
"""Plan- und Dokumentenliste Auflagebereinigung 2619-KISPI, Lenggstrasse 30.
Quelle: Vorabzug Bauentscheid (BE), Stadt Zürich. Eine Zeile pro Dokument/Plan,
gruppiert nach Planer, mit Frist-Spalte. JANS-Layout (neutral, schwarz)."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------ Stammdaten
PLANER = {
    "BRA": ("BRA", "Brandschutzexperte", "Jens Ziegel", "Gruner AG"),
    "HLK": ("HLK", "Heizung-, Lüftung-, Kälteplaner", "Cekdar Duran", "Gruner AG"),
    "SAN": ("SAN", "Sanitärplaner", "Mustafa Eren", "Gruner AG"),
    "SPR": ("SPR", "Sprinklerplaner", "Tim Harder", "JOMOS Brandschutz AG"),
    "ELE": ("ELE", "Elektroplaner", "Yasarcan Cetin", "Gruner AG"),
    "FKO": ("FKO", "Fachkoordination", "Benjamin Kübler", "Gruner AG"),
    "ARC": ("ARC", "Architekt", "Raphael Jans", "Raphael Jans Architekten ETH"),
    "BAU": ("BAU", "Bauleitung", "Albin Spahic", "KISPI"),
}
ORDER = ["BRA", "HLK", "SAN", "SPR", "ELE", "FKO", "ARC", "BAU"]

# Datenzeilen: (planer, dokument, auflagepunkt, fuer_amt, frist)
ROWS = {
"BRA": [
 ("Überarbeitetes Brandschutzkonzept (gesamter Umbaubereich)", "Ziff. II.16 / II.1b – BSK-Genehmigung (FP-Bestätigung an AfB)", "FP", "vor Baubeginn"),
 ("Überarbeitete Brandschutzpläne (gesamter Umbaubereich)", "Ziff. II.16 / II.1b", "FP", "vor Baubeginn"),
 ("Brandabschnittsplan EI30 + Türenliste EI30 (Patienten-/Personalzimmer, betriebstechn./Untersuchungs-/Behandlungs-/Technikräume)", "Ziff. II.14 – Brandabschnitte EI30", "FP", "vor Baubeginn"),
 ("Detailpläne Schiebetüren (Fluchttür + Brandabschluss) mit VKF-Anerkennung / Leistungserklärung, Beschrieb + Produktangaben", "Ziff. II.18 / II.1b – SRZ*EW1", "FP", "vor Baubeginn"),
 ("Detailpläne brandschutzrelevante Innenwandkonstruktionen (Materialangaben + Aufbau)", "Ziff. II.19 / II.1b", "FP", "vor Baubeginn"),
 ("Schott-/Abschottungskonzept Leitungsdurchführungen (RF1 / EI30, VKF-anerkannte Systeme)", "Ziff. II.20 – Abschottungen", "FP", "vor Ausführungsbeginn"),
 ("Brandfallsteuerungs-Konzept / Brandfallmatrix (auf BSK abgestimmt, dokumentiert + geprüft)", "Ziff. II.23 – Brandfallsteuerungen", "FP", "vor Inbetriebnahme"),
 ("Integral-Test-Protokolle Brandfallsteuerungen + Konformitätsbestätigung (VKF-BFS)", "Ziff. II.24 / II.2", "FP", "vor Inbetriebnahme"),
 ("Übereinstimmungserklärung Brandschutz (unterzeichnet durch QS-Verantwortliche/n)", "Ziff. II.2 – Inbetriebnahme", "FP", "vor Inbetriebnahme"),
 ("Nachgeführte Brandschutzpläne + Feuerwehrpläne (1. OG komplett) + nachgeführtes BSK (Revisionsunterlagen)", "Ziff. II.2", "FP", "vor Inbetriebnahme"),
],
"HLK": [
 ("Lüftungsplanung Therapieräume 1. OG (mechanische Lüftung nicht natürlich belüftbarer Räume)", "Ziff. II.4 – Lüftung Therapiestation", "AfB", "Ausführungsplanung"),
 ("Nachweis Fortluftführung über Dach (bei abweichender Lösung: Zustimmung einholen)", "Ziff. II.5 – Fortluft über Dach", "UGZ / StB", "Ausführungsplanung"),
 ("Nachweis Aussenluft-Ansaugöffnung ≥ 3 m über Terrain", "Ziff. II.6 – Aussenluftansaugung", "AfB / UGZ", "Ausführungsplanung"),
 ("Abgeänderte Pläne zusätzliche Lüftungskomponenten / Klima-Anlageteile an Fassade/Dach (ggf. mit Lärmgutachten)", "Ziff. II.7 – Zusatzkomponenten aussen", "AfB (Absprache UGZ/StB/DP)", "vor Erstellung"),
 ("Vollständiges Lüftungskonzept lufttechnische Anlagen (Umbaubereich), inkl. Reinigungsöffnungen + Brandschutzisolationen", "Ziff. II.21 / II.3", "FP", "vor Arbeitsvergabe"),
 ("Nachweis Anpassung lufttechnische Anlagen gemäss VKF-RL 'Lufttechnische Anlagen'", "Ziff. II.22", "FP", "vor Ausführungsbeginn"),
 ("Energienachweis – Hauptformular EN-ZH", "Ziff. II.8a / II.1c", "UGZ", "4 Wo vor Baubeginn"),
 ("Energienachweis – Lüftungstechnische Anlagen", "Ziff. II.8b", "UGZ", "4 Wo vor Baubeginn"),
 ("Energienachweis – Kühlung / Befeuchtung", "Ziff. II.8c", "UGZ", "4 Wo vor Baubeginn"),
 ("Nachweis genehmigte Abwärmenutzungsanlage Klima (oder Unwirtschaftlichkeitsnachweis, §30a BBV I)", "Erwägung d) / Ziff. II.8", "UGZ", "4 Wo vor Baubeginn"),
 ("Nachweis Klimatisierungsanforderungen SIA 180 / 382-1 (g-Wert + Sonnenschutz, Dichtigkeit, Wärmespeicher ≥ 30 Wh/m²K, Nutzungstrennung, kein gleichzeitiges Heizen/Kühlen)", "Erwägung e) / Ziff. II.8", "UGZ", "Ausführungsplanung"),
],
"SAN": [
 ("Abschottungen Sanitärleitungen durch Brandabschnitte (RF1/EI30) – Beitrag zum Schottkonzept", "Ziff. II.20", "FP (via BRA)", "vor Ausführungsbeginn"),
],
"SPR": [
 ("Anpassung Sprinkleranlage gemäss VKF-RL + Weisung 'Sprinkleranlagen' GVZ + Stand der Technik; Projektunterlagen einreichen + genehmigen lassen", "Ziff. II.26 / II.3", "GVZ", "vor Arbeitsvergabe"),
],
"ELE": [
 ("Anpassung Brandmeldeanlage gemäss VKF-RL + Weisung 'Brandmeldeanlagen' GVZ + Stand der Technik; Projektunterlagen einreichen + genehmigen lassen", "Ziff. II.25 / II.3", "GVZ", "vor Arbeitsvergabe"),
 ("Elektroseitige Umsetzung Brandfallsteuerungen (Ansteuerung/Verkabelung) – Beitrag Integral-Test", "Ziff. II.23 / II.24", "FP (via FKO)", "vor Inbetriebnahme"),
 ("Abschottungen Elektroleitungen durch Brandabschnitte (RF1/EI30) – Beitrag zum Schottkonzept", "Ziff. II.20", "FP (via BRA)", "vor Ausführungsbeginn"),
],
"FKO": [
 ("Koordination + Durchführung Integral-Test Brandfallsteuerungen (gewerkeübergreifend, Testorganisation + Protokollierung)", "Ziff. II.24 / II.2", "FP", "vor Inbetriebnahme"),
 ("Schnittstellen-Koordination gesamtheitliches Brandschutzkonzept (Übergänge zu Nachbarbereichen)", "Erwägung f)", "FP", "laufend"),
 ("Koordination Schott-/Abschottungskonzept zwischen HLK / SAN / ELE", "Ziff. II.20", "FP", "vor Ausführungsbeginn"),
],
"ARC": [
 ("Abgeänderte Pläne Fluchtwegführung (Fluchtwegdistanzen ≤ 20 m, Reduit 4.2 / Austauschpläne) – einreichen + bewilligen lassen", "Ziff. II.15 / II.1a – SRZ*EW2, Erwägung i)", "AfB", "vor Baubeginn"),
 ("Nachweis Hindernisfreiheit SIA 500:2009 (1. OG Therapiestation: betriebliche Anforderungen vorrangig)", "Erwägung b), §239a PBG", "AfB", "Ausführungsplanung"),
 ("Schallschutznachweis SIA 181 (Aussen-/Trennbauteile lärmempfindlicher Räume; haustechnische Anlagen mit HLK)", "Ziff. II.10 – Schallschutz", "UGZ (Lärm/NIS)", "Ausführungsplanung"),
 ("Ausführung gemäss bewilligten Plänen; Abweichungen vorgängig bewilligen lassen", "Ziff. II.33", "AfB", "laufend"),
 ("Kenntnisnahme Belästigungsverbot (Betrieb Gebäude / technische Anlagen) – kein separates Dokument", "Ziff. II.9", "—", "laufend"),
],
"BAU": [
 ("Anmeldung Bauinstallation / Bauzufahrt – Baukontrolle bei Benützung Privatgrund (Tel. 044 412 28 65)", "Ziff. II.28a", "AfB (Baukontrolle)", "4 Wo vor Baubeginn"),
 ("Anmeldung Benützung öffentlicher Grund – Bewilligung Gewerbe", "Ziff. II.28b", "StaPo", "4 Wo vor Baubeginn"),
 ("Abklärung Baumbestand (Wurzel-/Kronenbereich) – Freiraumberatung Baumerhalt", "Ziff. II.28c", "GSZ", "4 Wo vor Baubeginn"),
 ("Gemeinsames Zustandsprotokoll öffentlicher Grund + Festsetzung Depositum", "Ziff. II.29a/b", "TBA (Strassen)", "4 Wo vor Baubeginn"),
 ("Massnahmen Luftreinhaltung Baustelle (Massnahmenstufe A, Merkblatt Baustellen-Luft)", "Ziff. II.30", "UGZ / Vollzug", "während Bauzeit"),
 ("Meldung Baubeginn + Bauvollendung (§327 PBG, §23 BVV)", "Ziff. II.31", "AfB", "4 Wo vor / nach"),
 ("Bauabfall-Trennung + umweltgerechte Entsorgung (VVEA Art. 17), emissionsarme Transporte", "Ziff. II.32", "Vollzug (UGZ)", "während Bauzeit"),
 ("Organisatorischer Brandschutz Bauzeit (VKF-RL 'Brandverhütung / org. Brandschutz'), Flucht-/Rettungswege frei halten", "Ziff. II.27 / II.17", "FP / Vollzug", "während Bauzeit"),
],
}

# ------------------------------------------------------------------ Styling
BLACK = "FF000000"
HDR_FILL = PatternFill("solid", fgColor="FF404040")
GRP_FILL = PatternFill("solid", fgColor="FFD9D9D9")
ALT_FILL = PatternFill("solid", fgColor="FFF2F2F2")
TITLE_F  = Font(name="Cambria", size=14, bold=True, color=BLACK)
SUB_F    = Font(name="Cambria", size=10, color=BLACK)
HDR_F    = Font(name="Cambria", size=10, bold=True, color="FFFFFFFF")
GRP_F    = Font(name="Cambria", size=10, bold=True, color=BLACK)
CELL_F   = Font(name="Cambria", size=10, color=BLACK)
LEG_H    = Font(name="Cambria", size=10, bold=True, color=BLACK)
thin = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
TOP = Alignment(vertical="top", wrap_text=True)
TOPL = Alignment(vertical="top", horizontal="left", wrap_text=True)
CTR = Alignment(vertical="center", horizontal="left", wrap_text=True)

wb = Workbook()
ws = wb.active
ws.title = "Auflagebereinigung"
ws.sheet_view.showGridLines = False

# Spalten: Nr | Planer | Plan-/Dokument | Auflagepunkt | für Amt | Frist | Status
HEAD = ["Nr.", "Planer", "Plan- / Dokument", "Auflagepunkt", "für Amt", "Frist", "Status"]
WIDTHS = [5, 17, 46, 33, 18, 17, 14]
for i, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
NCOL = len(HEAD)

# Titelblock
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
c = ws.cell(1, 1, "Plan- und Dokumentenliste Auflagebereinigung")
c.font = TITLE_F; c.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

subs = [
 "Projekt 2619-KISPI «Neue Psychosomatische Therapiestation für Jugendliche» – "
 "Universitäts-Kinderspital Zürich, Eleonorenstiftung, Lenggstrasse 30, 8008 Zürich",
 "Grundlage: Vorabzug Bauentscheid (BE) Stadt Zürich  ·  kant. Beurteilung Verfügung BVV 26-1211, "
 "Baudirektion Kt. Zürich vom 18.05.2026 (kantonaler Betrieb, ohne Auflagen)",
 "Stand: 04.06.2026  ·  Status «Vorabzug» – noch nicht rechtskräftig  ·  Gliederung nach Planer "
 "gemäss GRUNER-Adressliste 2619 PPTS KISPI",
]
r = 2
for s in subs:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    cc = ws.cell(r, 1, s); cc.font = SUB_F; cc.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 26
    r += 1
r += 0
hdr_row = r + 1  # eine Leerzeile
# Header
for j, h in enumerate(HEAD, start=1):
    cc = ws.cell(hdr_row, j, h)
    cc.font = HDR_F; cc.fill = HDR_FILL; cc.alignment = CTR; cc.border = BORDER
ws.row_dimensions[hdr_row].height = 18

# Datenzeilen
rr = hdr_row + 1
nr = 0
data_first = rr
for code in ORDER:
    pc, role, name, firma = PLANER[code]
    # Gruppen-Kopf
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NCOL)
    gc = ws.cell(rr, 1, f"{pc}  ·  {role}  ·  {name} ({firma})")
    gc.font = GRP_F; gc.fill = GRP_FILL; gc.alignment = Alignment(vertical="center", horizontal="left")
    for j in range(1, NCOL + 1):
        ws.cell(rr, j).fill = GRP_FILL; ws.cell(rr, j).border = BORDER
    ws.row_dimensions[rr].height = 16
    rr += 1
    for k, (dok, aufl, amt, frist) in enumerate(ROWS[code]):
        nr += 1
        vals = [nr, pc, dok, aufl, amt, frist, "offen"]
        for j, v in enumerate(vals, start=1):
            cc = ws.cell(rr, j, v)
            cc.font = CELL_F; cc.border = BORDER
            if j in (1, 2, 5, 6, 7):
                cc.alignment = Alignment(vertical="top", horizontal="center" if j == 1 else "left", wrap_text=True)
            else:
                cc.alignment = TOPL
            if k % 2 == 1:
                cc.fill = ALT_FILL
        rr += 1
data_last = rr - 1

# Autofilter + Freeze
ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(NCOL)}{data_last}"
ws.freeze_panes = f"A{hdr_row + 1}"

# ------------------------------------------------------------------ Legenden
rr += 1
def legend_block(title, items):
    global rr
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NCOL)
    cc = ws.cell(rr, 1, title); cc.font = LEG_H
    rr += 1
    for it in items:
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NCOL)
        cc = ws.cell(rr, 1, it); cc.font = CELL_F; cc.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1
    rr += 1

legend_block("Ämter-Legende", [
 "AfB = Amt für Baubewilligungen Stadt Zürich   ·   FP = Feuerpolizei Stadt Zürich   ·   "
 "GVZ = Gebäudeversicherung Kanton Zürich",
 "UGZ = Umwelt- und Gesundheitsschutz Zürich, Energie im Bau   ·   StB = Amt für Städtebau   ·   "
 "DP = Denkmalpflege",
 "TBA = Tiefbauamt Stadt Zürich   ·   StaPo = Stadtpolizei Zürich   ·   GSZ = Grün Stadt Zürich",
])
legend_block("Fristen-Legende (Einreichungszeitpunkt gemäss Ziff. II.1–3)", [
 "Ziff. II.1 (vor Baubeginn, an AfB): a) Fluchtwegführung II.15  ·  b) Bestätigung FP zu II.16/II.18/II.19  ·  "
 "c) Bestätigung UGZ zu II.8 (energetisch)",
 "Ziff. II.2 (vor Inbetriebnahme / feuerpolizeil. Abnahme, an FP): Übereinstimmungserklärung, nachgeführte BS-/"
 "Feuerwehrpläne (1. OG), nachgef. BSK, Integral-Test-Protokolle",
 "Ziff. II.3 (vor Arbeitsvergabe): II.21 Lüftung  ·  II.25 BMA  ·  II.26 Sprinkler (SPA)",
 "4 Wo vor Baubeginn: Energienachweis-Formulare an UGZ (II.8)   ·   während Bauzeit: organisatorische Auflagen",
])
legend_block("Hinweise", [
 "• Eine Zeile = ein Deliverable (Plan/Dokument/Nachweis). Mehrere Zeilen können auf dieselbe Auflage-Ziffer zeigen.",
 "• SRZ*EW1 = Schiebetüren in brandabschnittsbildenden Bauteilen mit Fluchtwegfunktion (Erwägung h).  "
 "SRZ*EW2 = Reduit 4.2 (Lager 8 m²) nicht untergeordnet → Austauschpläne (Erwägung i).",
 "• «via BRA / via FKO» = Gewerk liefert den Fachbeitrag, federführende Einreichung über Brandschutz bzw. Fachkoordination.",
 "• Adress-Hinweis: offizielle Bauherrenadresse «Lenggstrasse 30» (GRUNER-Adressliste); Bauentscheid-Titel nennt «Lengstrasse 30» – bei Behörde verifizieren.",
 "• Status-Werte: offen / in Arbeit / eingereicht / genehmigt / erledigt / hinfällig.",
])
legend_block("Planerverzeichnis (Kontakte)", [
 "BRA  Jens Ziegel · Gruner AG · +41 43 299 74 32 · jens.ziegel@gruner.ch",
 "HLK  Cekdar Duran · Gruner AG · +41 61 367 96 46 · cekdar.duran@gruner.ch",
 "SAN  Mustafa Eren · Gruner AG · +41 61 367 95 96 · mustafa.eren@gruner.ch",
 "SPR  Tim Harder · JOMOS Brandschutz AG, Sagmattstrasse 5, CH-4710 Balsthal · "
 "+41 62 386 17 61 (Zentrale +41 62 386 17 17), Mobile +41 79 571 61 08 · tim.harder@jomos.ch",
 "ELE  Yasarcan Cetin · Gruner AG · +41 61 367 96 34 · yasarcan.cetin@gruner.ch",
 "FKO  Benjamin Kübler · Gruner AG · +41 61 367 96 44 · benjamin.kuebler@gruner.ch",
 "ARC  Raphael Jans · Raphael Jans Architekten ETH, Grubenstrasse 37, 8045 Zürich · +41 79 846 11 65 · rj@raphaeljans.ch",
 "BAU  Albin Spahic · Universitäts-Kinderspital Zürich (Leiter Technischer Dienst) · "
 "+41 44 249 31 50, Mobile +41 76 838 31 50 · albin.spahic@kispi.uzh.ch",
])

# ------------------------------------------------------------------ Speichern
OUT_PROJEKT = ("/Users/raphaeljans/Library/CloudStorage/OneDrive-FreigegebeneBibliotheken–JANS/"
 "JANS - 2619-KISPI - Dokumente/3 Umbauprojekt Neu/33.01_JANS/01 Dokumente Auflagebereinigung/"
 "_Checklisten Architekt Plan und Dokumentenliste Komplette Auflagen Insgesamt")
OUT_HUB = ("/Users/raphaeljans/Library/CloudStorage/OneDrive-FreigegebeneBibliotheken–JANS/"
 "AD - 01 Geschaeftsfuerung/JANS AI/30 JANS AI HUB OUTPUT/auflagebereinigung/2619-KISPI")
FNAME = "260604-Plan-und-Dokumentenliste-Auflagebereinigung-2619-KISPI.xlsx"

os.makedirs(OUT_HUB, exist_ok=True)
p1 = os.path.join(OUT_PROJEKT, FNAME)
p2 = os.path.join(OUT_HUB, FNAME)
wb.save(p1)
wb.save(p2)
print("Zeilen (Deliverables):", nr)
print("gespeichert:", p1)
print("gespeichert:", p2)
