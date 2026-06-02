#!/usr/bin/env python3
"""
Vergleichsmatrix KISPI Therapiestation Gastrokueche
Rametall AG (Plan .3) vs. Gastro-Online AG (Plan .4)
Ausgabe: A4 Querformat, eine Seite, klar lesbar.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PrintOptions, PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

OUT = '/Users/raphaeljans/Library/CloudStorage/OneDrive-FreigegebeneBibliotheken–JANS/AD - 01 Geschaeftsfuerung/JANS AI/30 JANS AI HUB OUTPUT/submission/2619-kispi-gastrokueche/auswertung/260513_Vergleichsmatrix_KISPI_Gastrokueche.xlsx'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Vergleich"

# === Styles ===
font_title = Font(name='Cambria', size=14, bold=True, color='000000')
font_sub   = Font(name='Cambria', size=9, italic=True, color='000000')
font_head  = Font(name='Cambria', size=9, bold=True, color='000000')
font_body  = Font(name='Cambria', size=9, color='000000')
font_bold  = Font(name='Cambria', size=9, bold=True, color='000000')
font_total = Font(name='Cambria', size=10, bold=True, color='000000')
font_small = Font(name='Cambria', size=8, color='000000')

thin = Side(border_style='thin', color='000000')
hairline = Side(border_style='hair', color='808080')
no_border = Side(border_style=None)

border_top    = Border(top=thin, bottom=no_border, left=no_border, right=no_border)
border_bottom = Border(top=no_border, bottom=thin, left=no_border, right=no_border)
border_both   = Border(top=thin, bottom=thin, left=no_border, right=no_border)
border_row    = Border(top=no_border, bottom=hairline, left=no_border, right=no_border)

fill_section = PatternFill('solid', fgColor='F2F2F2')
fill_total   = PatternFill('solid', fgColor='E8E8E8')

al_left  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
al_right = Alignment(horizontal='right',  vertical='center', wrap_text=True)
al_ctr   = Alignment(horizontal='center', vertical='center', wrap_text=True)

# === Seite (A4 quer) ===
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize   = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth  = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4, header=0.2, footer=0.2)
ws.print_options = PrintOptions(horizontalCentered=True)

# === Spaltenbreiten (A4 quer 27 cm Inhaltsbreite) ===
# Spalten: Pos | Bezeichnung | Menge | Rametall | Gastro | Differenz | Status
widths = {'A': 6, 'B': 48, 'C': 5, 'D': 13, 'E': 13, 'F': 13, 'G': 28}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# === Titel ===
ws.merge_cells('A1:G1')
ws['A1'] = "Vergleichsmatrix Kuecheneinrichtung — KISPI Therapiestation"
ws['A1'].font = font_title
ws['A1'].alignment = al_left
ws.row_dimensions[1].height = 22

ws.merge_cells('A2:G2')
ws['A2'] = "Rametall AG (Plan 26-122-01.3, 6.5.2026) vs. Gastro-Online AG (Plan 26-122-01.4, 13.5.2026)  -  Alle Preise CHF exkl. MwSt  -  Stand 13. Mai 2026"
ws['A2'].font = font_sub
ws['A2'].alignment = al_left
ws.row_dimensions[2].height = 14

# === Tabellen-Header ===
headers = ['Pos.', 'Bezeichnung / Spezifikation', 'Mge', 'Rametall AG', 'Gastro-Online AG', 'Differenz', 'Status / Bemerkung']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = font_head
    c.alignment = al_left if i in (1,2,7) else (al_ctr if i == 3 else al_right)
    c.border = border_both
ws.row_dimensions[4].height = 18

# === Datenzeilen ===
rows = []

def section(title):
    rows.append(('SECTION', title))

def pos(pos, bez, mge, r, g, diff, status):
    rows.append(('POS', pos, bez, mge, r, g, diff, status))

def subtotal(label, r, g):
    rows.append(('SUB', label, r, g))

# ===== 1.0 Spuelen / Ruesten =====
section("1.0 Spuelen / Ruesten")
pos('1.01', 'Arbeitstisch 3380 x 900/730 mit Unterbau (Fluegeltuere, Abfallfach, Nische GSP, Doppelfluegeltuere)', 1, 7900, 9830, -1930, 'BEIDE  -  Gastro +24%')
pos('1.02', 'Untertisch-Geschirrspuelmaschine Meiko Upster', 1, 'bestehend', 'bestehend', '—', 'BESTEHEND beide')
pos('1.03', 'Servicewagen 3-bordig', 1, '450 (CNS)', 'bestehend (Kunststoff schwarz)', '—', 'ANDERS  -  Gastro nimmt Bestand')
pos('1.04', 'Gemueseschneidemaschine Rotor Varimat Speed Gourmet, 5 Schneidescheiben', 1, 4950, '—', '+4950 R', 'FEHLT bei Gastro')
pos('1.05', 'Wandhaengeschrank Fluegeltueren 800 x 360 x 760  (1 von 2)', 1, 930, 1250, -320, 'BEIDE  -  Gastro +34%')
pos('1.06', 'Wandhaengeschrank Fluegeltueren 800 x 360 x 760  (2 von 2)', 1, 930, 1250, -320, 'BEIDE  -  Gastro +34%')
subtotal('Zwischensumme 1.0', 15160, 12330)

# ===== 2.0 Warme Kueche =====
section("2.0 Warme Kueche")
pos('2.01', 'Arbeitstisch warme Kueche mit Becken 340 x 370 + Kuehlunterbau  -  Plan .3: 2289 mm, 2x GN-1/1  /  Plan .4: 3040 mm, 1x GN-1/1 + Nische UT-TKS', 1, 6130, 14300, -8170, 'PLAN UNTERSCHIEDLICH (.3 vs .4)')
pos('2.02', 'Rational iCombi Pro 6 x GN-2/3, UltraVent Plus, Tischuntergestell, 6 Roste', 1, 11900, 12100, -200, 'BEIDE  -  Markengleich, +1.7%')
pos('2.03G', 'Untertisch-Tiefkuehlschrank EK 595 x 642 x 830', 1, 'bestehend', 'bestehend', '—', 'BESTEHEND  -  nur Plan .4 separat gefuehrt')
pos('2.04 / 2.03', 'Wandhaengeschrank warme Kueche  -  Plan .3: Fluegeltueren 957 mm  /  Plan .4: Schiebetueren 1200 mm', 1, 1040, 1500, -460, 'ANDERS  -  Gastro groesser, Schiebetueren')
pos('2.05G', 'Wandhaengeschrank offen 1000 x 360 x 760 mit Tablar', 1, '—', 670, '-670 G', 'FEHLT bei Rametall (nur Plan .4)')
pos('2.04R / 4.01G', 'Regalwagen mit Arbeitsplatte, 8 x GN-2/1 und Lenkrollen', 1, 830, 970, -140, 'BEIDE  -  bei Gastro in 4.0')
subtotal('Zwischensumme 2.0', 19900, 28570)

# ===== 3.0 Herdanlage =====
section("3.0 Herdanlage")
pos('3.01', 'Herd / Korpus Kochen  -  Plan .3 (Rametall handschr. korrigiert): Infrarot 2 x 3 kW, total 12 kW, Korpus 2100 x 850, Glas Berner BS2ZES  /  Plan .4 (Gastro): Power-Induktion 2 x 2er a 360 x 660 mm, Korpus 3300 x 1100, Tablarfaecher, Induktionsfuss', 1, 12940, 26750, -13810, 'UNTERANGEBOT Rametall  -  Funktion!')
pos('3.02', 'Ablufthaube ca. 1200 x 1200 x 450', 1, 'bauseits', 'bauseits', '—', 'BAUSEITS beide (HLK)')
subtotal('Zwischensumme 3.0', 12940, 26750)

# ===== 4.0 Vorbereiten =====
section("4.0 Vorbereiten")
pos('4.01', 'Arbeitstisch 1200 x 600 x 900 mm fahrbar mit Korpus 3 Schubladen, Tablarfach offen, 4 Lenkrollen', 1, 3350, '—', '+3350 R', 'FEHLT bei Gastro')
pos('4.02', 'Arbeitstisch 1000 x 600 x 900 mm fahrbar mit 2 Fluegeltuerfaechern, 4 Lenkrollen', 1, 2680, '—', '+2680 R', 'FEHLT bei Gastro')
pos('4.03', 'Servicewagen 3-bordig, CNS-Ausfuehrung', 1, 450, '—', '+450 R', 'FEHLT bei Gastro')
subtotal('Zwischensumme 4.0', 6480, 0)

# ===== 5.0 Lager / Economat =====
section("5.0 Lager / Economat")
pos('5.01', 'Doppeltuerkuehlschrank GN-2/1, ca. 1400 l, eigengekuehlt', 1, 'bestehend', 'bestehend', '—', 'BESTEHEND beide')
pos('5.02', 'Kuehlkorpus mit 2 Kuehlschubladen, eigengekuehlt, CNS-Arbeitsplatte', 1, 'bestehend', 'bestehend', '—', 'BESTEHEND beide')
pos('5.03R / 6.02G', 'CNS-Lagerschrank Fluegeltueren 1000 x 590 x 2000, 4 Tablare, Sockel', 1, 2330, 'bauseits', '+2330 R', 'BAUSEITS bei Gastro')
pos('5.04R / 6.04G', 'ScanBox Warmhaltewagen', 1, 'bauseits', 'bauseits', '—', 'BAUSEITS beide')
pos('5.05R / 5.04G', 'Aluminiumregal 4-lagig mit PE-Auflagen 1200 x 460 x 1675', 1, 630, 480, +150, 'BEIDE  -  Rametall +31%')
pos('5.06R / 5.06G', 'Aluminiumregal 4-lagig mit PE-Auflagen 770 x 460 x 1675  (1 von 2 R)', 1, 470, 350, +120, 'BEIDE  -  Rametall +34%')
pos('5.03G',         'Aluminiumregal 4-lagig 770 x 460 x 1675  (zusaetzlich Gastro)', 1, '—', 350, '-350 G', 'FEHLT bei Rametall')
pos('5.05G',         'Aluminiumregal 4-lagig 770 x 460 x 1675  (zusaetzlich Gastro)', 1, '—', 350, '-350 G', 'FEHLT bei Rametall')
subtotal('Zwischensumme 5.0', 3430, 1530)

# ===== 6.0 Lager =====
section("6.0 Lager")
pos('6.01R / 6.03G', 'Aluminiumregal 4-lagig 770 x 460 x 1675  (2 von 2 R)', 1, 470, 350, +120, 'BEIDE  -  Rametall +34%')
pos('6.02R / 6.01G', 'Putzschrank 600 x 590 x 2000, abschliessbar, mit Tablaren, Putzmittelkorb, Besenhalterung', 1, 2180, 'bauseits', '+2180 R', 'BAUSEITS bei Gastro')
subtotal('Zwischensumme 6.0', 2650, 350)

# ===== Schreiben =====
row = 5
for r in rows:
    if r[0] == 'SECTION':
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=r[1])
        c.font = font_bold
        c.alignment = al_left
        c.fill = fill_section
        c.border = border_bottom
        ws.row_dimensions[row].height = 14
    elif r[0] == 'POS':
        _, p, bez, mge, rr, gg, diff, status = r
        ws.cell(row=row, column=1, value=p).font = font_body
        ws.cell(row=row, column=2, value=bez).font = font_body
        ws.cell(row=row, column=3, value=mge).font = font_body
        ws.cell(row=row, column=4, value=rr).font = font_body
        ws.cell(row=row, column=5, value=gg).font = font_body
        ws.cell(row=row, column=6, value=diff).font = font_body
        ws.cell(row=row, column=7, value=status).font = font_body
        ws.cell(row=row, column=1).alignment = al_left
        ws.cell(row=row, column=2).alignment = al_left
        ws.cell(row=row, column=3).alignment = al_ctr
        ws.cell(row=row, column=4).alignment = al_right
        ws.cell(row=row, column=5).alignment = al_right
        ws.cell(row=row, column=6).alignment = al_right
        ws.cell(row=row, column=7).alignment = al_left
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border_row
        if isinstance(diff, (int, float)) and diff != 0:
            ws.cell(row=row, column=6).number_format = '#,##0;-#,##0'
        if isinstance(rr, (int, float)):
            ws.cell(row=row, column=4).number_format = '#,##0'
        if isinstance(gg, (int, float)):
            ws.cell(row=row, column=5).number_format = '#,##0'
        ws.row_dimensions[row].height = 22
    elif r[0] == 'SUB':
        _, label, rr, gg = r
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=label)
        c.font = font_bold
        c.alignment = al_right
        c.fill = fill_total
        ws.cell(row=row, column=4, value=rr).font = font_bold
        ws.cell(row=row, column=4).alignment = al_right
        ws.cell(row=row, column=4).fill = fill_total
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=5, value=gg).font = font_bold
        ws.cell(row=row, column=5).alignment = al_right
        ws.cell(row=row, column=5).fill = fill_total
        ws.cell(row=row, column=5).number_format = '#,##0'
        diff = rr - gg
        ws.cell(row=row, column=6, value=diff).font = font_bold
        ws.cell(row=row, column=6).alignment = al_right
        ws.cell(row=row, column=6).fill = fill_total
        ws.cell(row=row, column=6).number_format = '#,##0;-#,##0'
        ws.cell(row=row, column=7, value='').fill = fill_total
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border_both
        ws.row_dimensions[row].height = 16
    row += 1

# ===== Total Block =====
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value='Summe der Zwischensummen (rechnerisch)').font = font_bold
ws.cell(row=row, column=1).alignment = al_right
ws.cell(row=row, column=4, value=60160).font = font_bold
ws.cell(row=row, column=4).alignment = al_right
ws.cell(row=row, column=4).number_format = '#,##0'
ws.cell(row=row, column=5, value=70500).font = font_bold
ws.cell(row=row, column=5).alignment = al_right
ws.cell(row=row, column=5).number_format = '#,##0'
ws.cell(row=row, column=6, value=-10340).font = font_bold
ws.cell(row=row, column=6).alignment = al_right
ws.cell(row=row, column=6).number_format = '#,##0;-#,##0'
ws.cell(row=row, column=7, value='Rametall klar guenstiger').font = font_bold
ws.cell(row=row, column=7).alignment = al_left
for col in range(1, 8):
    ws.cell(row=row, column=col).border = border_top
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value='Ausgewiesene Summe Kuecheneinrichtung').font = font_body
ws.cell(row=row, column=1).alignment = al_right
ws.cell(row=row, column=4, value=64160).font = font_body
ws.cell(row=row, column=4).alignment = al_right
ws.cell(row=row, column=4).number_format = '#,##0'
ws.cell(row=row, column=5, value=70500).font = font_body
ws.cell(row=row, column=5).alignment = al_right
ws.cell(row=row, column=5).number_format = '#,##0'
ws.cell(row=row, column=7, value='Rametall: +4000 Additionsfehler').font = font_small
ws.cell(row=row, column=7).alignment = al_left
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value='Projektrabatt (nur Gastro)').font = font_body
ws.cell(row=row, column=1).alignment = al_right
ws.cell(row=row, column=4, value='—').font = font_body
ws.cell(row=row, column=4).alignment = al_right
ws.cell(row=row, column=5, value=-2800).font = font_body
ws.cell(row=row, column=5).alignment = al_right
ws.cell(row=row, column=5).number_format = '#,##0;-#,##0'
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value='Lieferung / Montage / Inbetriebnahme').font = font_body
ws.cell(row=row, column=1).alignment = al_right
ws.cell(row=row, column=4, value=5000).font = font_body
ws.cell(row=row, column=4).alignment = al_right
ws.cell(row=row, column=4).number_format = '#,##0'
ws.cell(row=row, column=5, value=4700).font = font_body
ws.cell(row=row, column=5).alignment = al_right
ws.cell(row=row, column=5).number_format = '#,##0'
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1, value='Gesamttotal exkl. MwSt (gemaess Offerte)').font = font_total
ws.cell(row=row, column=1).alignment = al_right
ws.cell(row=row, column=1).fill = fill_total
ws.cell(row=row, column=4, value=69160).font = font_total
ws.cell(row=row, column=4).alignment = al_right
ws.cell(row=row, column=4).fill = fill_total
ws.cell(row=row, column=4).number_format = '#,##0'
ws.cell(row=row, column=5, value=72400).font = font_total
ws.cell(row=row, column=5).alignment = al_right
ws.cell(row=row, column=5).fill = fill_total
ws.cell(row=row, column=5).number_format = '#,##0'
ws.cell(row=row, column=6, value=-3240).font = font_total
ws.cell(row=row, column=6).alignment = al_right
ws.cell(row=row, column=6).fill = fill_total
ws.cell(row=row, column=6).number_format = '#,##0;-#,##0'
ws.cell(row=row, column=7, value='Rametall 4.5% guenstiger').font = font_total
ws.cell(row=row, column=7).alignment = al_left
ws.cell(row=row, column=7).fill = fill_total
for col in range(1, 8):
    ws.cell(row=row, column=col).border = border_both

# === Analyse-Block am Ende ===
row += 2
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
ws.cell(row=row, column=1, value='Analyse — was bedeutet das?').font = font_head
ws.cell(row=row, column=1).alignment = al_left
row += 1
analyse = [
    'A) Gleich offerierte Positionen (in beiden Offerten enthalten, gleiche Spezifikation):  Rametall 24 060 CHF  /  Gastro 26 580 CHF  →  Gastro +2 520 CHF (+10.5%) teurer.',
    'B) Nur in Rametall offeriert (bei Gastro fehlt oder bauseits):  Gemueseschneider 4950, Arbeitstisch 1200x600 fahrbar 3350, Arbeitstisch 1000x600 fahrbar 2680, Servicewagen CNS 450, CNS-Lagerschrank 2330, Putzschrank 2180  →  Summe Mehrleistung Rametall = 15 940 CHF.',
    'C) Nur in Gastro offeriert (bei Rametall fehlt):  Wandhaengeschrank offen 1000 mm 670, 2 zusaetzliche Aluregale 770 (700)  →  Summe Mehrleistung Gastro = 1 370 CHF.',
    'D) Spezifikations-Differenz: Herdanlage  -  Rametall liefert Infrarot 12 kW statt der im Plan .4 vorgesehenen Power-Induktion (UNTERANGEBOT um ca. 13 800 CHF; Plan-/Funktions-Abweichung).',
    'E) Bereinigte Schaetzung bei vollem Lieferumfang nach Plan .4:  Rametall 77 000 - 80 000 CHF  /  Gastro 88 000 - 91 000 CHF  →  Rametall ca. 10 000 - 13 000 CHF (12 - 15%) guenstiger.',
]
for line in analyse:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1, value=line)
    c.font = font_small
    c.alignment = al_left
    ws.row_dimensions[row].height = 22
    row += 1

# Footer
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
ws.cell(row=row, column=1, value='Raphael Jans Architekten ETH SIA  -  rj@raphaeljans.ch  -  13. Mai 2026').font = font_small
ws.cell(row=row, column=1).alignment = al_left

# Print area
ws.print_area = f'A1:G{row}'

wb.save(OUT)
print("xlsx geschrieben:", OUT)
